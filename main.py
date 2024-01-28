from dotenv import load_dotenv
import shutil
import asyncio
import sys
import os
import openpyxl
from openapp import (openCRM,connectToCRM)
from searchandupdateot import (searchAndUpdateDates,searchAndUpdateUser,searchAndUpdateBillingItem,searchAndUpdateState)
from closeapps import (closeCRM)
from genericfunctions import (showDesktop,make_noise,get_username_os,sendEmail,getCurrentDateAndTime,terminateProcess)
from telegramfunctions import (sendTelegramMsg,sendTelegramMsgWithDocuments)
#import logging
import menu
from datetime import datetime 
from time import sleep

def exit_program():
    sys.exit(0)

def reset_crm():    
    terminateProcess('CRM.exe')                
    sleep(2)
    openCRM()
    connectToCRM()

async def main():       

    # VARIABLES DE ENTORNO
    load_dotenv()        
    TELEGRAM_CHAT_ID = os.getenv('TELEGRAM_CHAT_ID') # PERSONAL CHAT WITH BOT 5970685607 # TELEGRAM CHAT GROUP ID #-1002019721248
    INPUT_DIRECTORY = os.getenv('INPUT_DIRECTORY')
    INPUT_FILENAME = os.getenv('INPUT_FILENAME')
    #SUPER_LOG_FILENAME = 'C:/BOT BPO Automation/Version 1.0/logs/super_log.txt'
    GENERIC_ERROR_MSG = 'No Se ha procesado el incidente - Para ver mas detalles ver el archivo de logs'    

    # VARIABLES        
    counter = 1
    total_rows = 0
    total_cols = 0
    resp_funcion = None
    
    try:
        #***************************************#
        #******* LOGGING CONFIGURATION *********#
        #***************************************#
        
        # SUPER LOGGER
        # FORMAT ='%(asctime)s - %(name)s - %(user)s - %(levelname)s - %(message)s'
        # logging.basicConfig(filename=SUPER_LOG_FILENAME, filemode='w',format=FORMAT, datefmt='%m/%d/%Y %I:%M:%S %p',level=logging.WARNING)
        # attrs = {'user': get_username_os()}
               
        #*******************************************#
        #*** SHOW MENU TO SELECT BOT ACTIVITY ******#
        #*******************************************#        
        act_bot = menu.show()
        
        #*******************************************#
        #************* DATA EXTRACTION *************#
        #*******************************************#
        # Extract Data from Excel File and Open an existing Excel wb file 
        
        file = os.path.join(INPUT_DIRECTORY,INPUT_FILENAME)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        if 'Input Backup' in wb.sheetnames:
            del wb['Input Backup']
        
        # Notify the start of BOT execution via Telegram                
        #await sendTelegramMsg('START - BOT ' + actividad_rpa_selected + ' ha iniciado.\nUsuario ['+ get_username_os() +']\nTiempo de Inicio: ' + getCurrentDateAndTime(),TELEGRAM_CHAT_ID)
        #await sendTelegramMsgWithDocuments(TELEGRAM_CHAT_ID)        
        # # Notify the start of BOT execution via Email         
        # sendEmail(
        #     'START - BOT ' + actividad_rpa_selected + ' ha iniciado',
        #     CORREOS_DESTINO,
        #     CORREOS_CC,
        #     'START - BOT ' + actividad_rpa_selected + ' ha sido iniciado por el usuario ['+ get_username_os() +'] a las: ' + getCurrentDateAndTime(),
        #     'C:\\BOT INSUMO BASE\\Input BOT 001 V1.xlsx',
        #     'null'            
        # )             
        
        # Excel Input Base Manipulation
        target_ws = wb.copy_worksheet(ws)
        target_ws.title = "Input Backup"
        total_rows = len(ws['A'])
        total_cols = len(ws[1])
        print("Total de registros en base fuente: ",total_rows-1)
        print("Total de columnas en base fuente: ",total_cols)
        #logging.warning('%s registros en base fuente',total_rows-1,extra=attrs)
        #logging.warning('Extract Data Process Finished',extra=attrs)        
        
        #*************************************#
        #************* OPEN APPS *************#
        #*************************************#
        
        #CRM        
        openCRM()
        sleep(1)
        connectToCRM()        
        
        #*******************************************#
        #************* DATA PROCESSING *************#
        #*******************************************#        
        #Looping with iter_rows method through all the OTs (registers in the Excel)

        rows = ws.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=total_cols)                    
        for row in rows:
            # se omiten aquellos registros que en la col STATUS tenga COMPLETADO o contenga la palabra ERROR            
            if row[1].value == 'COMPLETADO' or str(row[1].value).find("ERROR") != -1:
                continue

            #Interacting with the CRM Depending on the user input call the corresponding method            
            if act_bot == 1:
                print('Validando registro: ',str(row[0].value),row[3].value,row[4].value)                
                resp_funcion = searchAndUpdateDates(str(row[0].value),row[3].value,row[4].value,str(row[10].value))
                print("Respuesta de la función:",resp_funcion)
                
            if act_bot == 2:
                resp_funcion = searchAndUpdateUser(str(row[0].value),str(row[6].value))  
                print("Respuesta de la función:",resp_funcion)

            if act_bot == 3:
                resp_funcion = searchAndUpdateBillingItem(str(row[0].value),str(row[8].value))  
                print("Respuesta de la función:",resp_funcion)
                
            if act_bot == 4:
                resp_funcion = searchAndUpdateState(str(row[0].value),str(row[5].value),str(row[6].value),str(row[9].value))  
                print("Respuesta de la función:",resp_funcion)
            
            if act_bot == 5:
                resp_funcion = searchAndUpdateState(str(row[0].value),str(row[5].value),str(row[6].value),str(row[9].value))  
                print("Respuesta de la función:",resp_funcion)            
            
            # Manejo de las respuestas identificadas en los metodos de control en CRM
            if resp_funcion == 0:                
                row[1].value = 'COMPLETADO'
            elif resp_funcion == 10:
                row[1].value = 'ERROR - No se identifica vista de detalles de OT reconocida en el CRM'                
            elif resp_funcion == 9:                
                row[1].value = 'ERROR - Se presenta mensaje de advertencia en el proceso en CRM'                    
            else:
                row[1].value = 'ERROR - '+ GENERIC_ERROR_MSG                
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                sleep(2)
                openCRM()            
                connectToCRM()
                
            print("Registro (OT):",counter)
            if counter % 10 == 0:                
                await sendTelegramMsg('Se han procesado '+ str(counter) +' registros, reiniciando CRM... ',TELEGRAM_CHAT_ID)
                reset_crm()
                                                            
            counter = counter + 1
            wb.save(file)                        
            
        print("Fin del proceso principal")

    except FileNotFoundError as e:
        
        await sendTelegramMsg('END - Error de Ejecución: Archivo Fuente No Encontrado',TELEGRAM_CHAT_ID)
        # sendEmail(
        #     'END - Error de Ejecución: FileNotFoundError - BOT ' + actividad_rpa_selected,
        #     CORREOS_DESTINO,
        #     CORREOS_CC,
        #     'END - Ha ocurrido un error en la ejecución del ' + actividad_rpa_selected + '.\nNo se encuentra el archivo base fuente\n' + 'Usuario ['+ get_username_os() +']\nTiempo de finalización:'+ getCurrentDateAndTime(),
        #     'null',
        #     'null'
        # )
        exit_program()
        
    except Exception as e:
        
        print(f"An error occurred: {e}")
        print(sys.exc_info())        
        e_type, e_object, e_traceback = sys.exc_info()
        e_line_number = e_traceback.tb_lineno        
        await sendTelegramMsg('END - Error de Ejecución: ' + str(e) + '.\n'+'Tipo de Error: '+str(sys.exc_info()[0]),TELEGRAM_CHAT_ID)
        # sendEmail(
        #     'END - Error de Ejecución: ' + e + ' - BOT ' + actividad_rpa_selected,
        #     CORREOS_DESTINO,
        #     CORREOS_CC,           
        #     'END - Ha ocurrido un error en la ejecución del BOT actividad_rpa_selected.\nError de Ejecución: ' + e + '\nUsuario ['+ get_username_os() +']\nTiempo de finalización:'+ getCurrentDateAndTime(),
        #     'null',
        #     'null'
        # )
        exit_program()
    
    else:
        #*******************************************#
        #******** IF NO ERRORS OCCURRED ************#
        #*******************************************#
        
        #***************************************#
        #************* CLOSE APPS  *************#
        #***************************************#
        closeCRM()
        sleep(1)
        showDesktop()                            

        #***************************************#
        #*********** SEND NOTIFICATIONS  *******#
        #***************************************#
                
        #logging.warning('Main Function finished naturally',extra=attrs)
        #await sendTelegramMsg('END - BOT '+actividad_rpa_selected+' ha finalizado.\n' + 'Usuario ['+ get_username_os() +']\nTiempo de finalización:'+ getCurrentDateAndTime(),TELEGRAM_CHAT_ID)
        await sendTelegramMsg('END - BOT '+' ha finalizado.\n' + 'Usuario ['+ get_username_os() +']\nTiempo de finalización:'+ getCurrentDateAndTime(),TELEGRAM_CHAT_ID)
        await sendTelegramMsgWithDocuments(TELEGRAM_CHAT_ID)        
        # sendEmail(
        #     'END - BOT '+ actividad_rpa_selected +' ha finalizado',
        #     CORREOS_DESTINO,
        #     CORREOS_CC,
        #     'END - BOT ' + actividad_rpa_selected + ' ha finalizado. \n' + 'Usuario ['+ get_username_os() +']\n Tiempo de finalización:'+ getCurrentDateAndTime(),
        #     'C:\\BOT INSUMO BASE\\Input BOT 001 V1.xlsx',
        #     'null'
        # )
        exit_program()        
    finally:
        
        # if (os.path.exists("C:/BOT BPO Automation/Version 1.0/logs/super_log.txt") == False):
        #     f = open("C:/BOT BPO Automation/Version 1.0/logs/super_log.txt")
        #     f.close()
        # else:
        #     print("File Exists")         
        # shutil.copyfile('C:/BOT BPO Automation/Version 1.0/logs/super_log.txt', 'C:/BOT BPO Automation/Version 1.0/logs/old/super_log_'+ str(datetime.now()).translate(str.maketrans({':': '-', '.': '-'}))+ '.txt')
        # sleep(1)
        make_noise()     

if __name__ == "__main__":
   asyncio.run(main())